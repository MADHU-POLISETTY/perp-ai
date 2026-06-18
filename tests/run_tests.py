import os
import sys
import time
import unittest
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure workspace paths are in sys.path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

def clean_scenario_name(method_name):
    # Converts test_simulate_successful_login to Simulate Successful Login
    parts = method_name.split("_")
    if parts[0] == "test":
        parts = parts[1:]
    return " ".join(part.capitalize() for part in parts)

def get_component_name(method_name):
    method_lower = method_name.lower()
    if "auth" in method_lower:
        return "Authentication"
    elif "landing" in method_lower:
        return "Landing Page"
    elif "dashboard" in method_lower or "sidebar" in method_lower or "menu" in method_lower or "hamburger" in method_lower:
        return "Dashboard / Sidebar"
    elif "session" in method_lower or "question" in method_lower or "answer" in method_lower or "coaching" in method_lower:
        return "Interview Session"
    elif "results" in method_lower or "gauge" in method_lower:
        return "Evaluation Results"
    elif "history" in method_lower or "deletion" in method_lower or "logs" in method_lower:
        return "Histories & Logs"
    elif "resume" in method_lower:
        return "Resume Analyzer"
    elif "profile" in method_lower or "charts" in method_lower or "analytics" in method_lower:
        return "Candidate Analytics"
    else:
        return "General System"

def main():
    print("=" * 80)
    print("                    PREPWISE AI AUTOMATION TEST RUNNER")
    print("=" * 80)
    print("Discovering test cases...")

    # Discover and load tests
    loader = unittest.defaultTestLoader
    start_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir = os.path.dirname(start_dir)
    suite = loader.discover(start_dir, pattern="test_*.py", top_level_dir=root_dir)

    print(f"Total test suites discovered: {suite.countTestCases()} test cases.")
    print("Executing tests...")
    
    test_results = []
    
    # Custom test runner to extract execution results
    class CustomResult(unittest.TestResult):
        def startTest(self, test):
            self._start_time = time.time()
            super().startTest(test)

        def addSuccess(self, test):
            elapsed = time.time() - self._start_time
            method_name = test._testMethodName
            category = "Appium (Mobile)" if "appium" in test.__class__.__module__.lower() else "Selenium (Desktop)"
            scenario = test._testMethodDoc or clean_scenario_name(method_name)
            component = get_component_name(method_name)
            
            test_results.append({
                "id": f"TS-{len(test_results) + 1:03d}",
                "category": category,
                "component": component,
                "scenario": scenario.strip(),
                "status": "PASS",
                "time": round(elapsed, 3),
                "error": "",
                "screenshot": ""
            })
            super().addSuccess(test)

        def addFailure(self, test, err):
            elapsed = time.time() - self._start_time
            method_name = test._testMethodName
            category = "Appium (Mobile)" if "appium" in test.__class__.__module__.lower() else "Selenium (Desktop)"
            scenario = test._testMethodDoc or clean_scenario_name(method_name)
            component = get_component_name(method_name)
            error_details = self._exc_info_to_string(err, test)
            
            # Screenshot capture on failure
            driver = getattr(test, "driver", None) or getattr(test.__class__, "driver", None)
            screenshot_relative_path = ""
            if driver:
                try:
                    screenshots_dir = os.path.join(start_dir, "screenshots")
                    os.makedirs(screenshots_dir, exist_ok=True)
                    s_name = f"{test.__class__.__name__}_{method_name}.png"
                    s_path = os.path.join(screenshots_dir, s_name)
                    driver.save_screenshot(s_path)
                    screenshot_relative_path = f"screenshots/{s_name}"
                except Exception as e:
                    print(f"Failed to capture failure screenshot: {e}")

            test_results.append({
                "id": f"TS-{len(test_results) + 1:03d}",
                "category": category,
                "component": component,
                "scenario": scenario.strip(),
                "status": "FAIL",
                "time": round(elapsed, 3),
                "error": error_details,
                "screenshot": screenshot_relative_path
            })
            super().addFailure(test, err)

        def addError(self, test, err):
            elapsed = time.time() - self._start_time
            method_name = test._testMethodName
            category = "Appium (Mobile)" if "appium" in test.__class__.__module__.lower() else "Selenium (Desktop)"
            scenario = test._testMethodDoc or clean_scenario_name(method_name)
            component = get_component_name(method_name)
            error_details = self._exc_info_to_string(err, test)
            
            # Screenshot capture on error
            driver = getattr(test, "driver", None) or getattr(test.__class__, "driver", None)
            screenshot_relative_path = ""
            if driver:
                try:
                    screenshots_dir = os.path.join(start_dir, "screenshots")
                    os.makedirs(screenshots_dir, exist_ok=True)
                    s_name = f"{test.__class__.__name__}_{method_name}.png"
                    s_path = os.path.join(screenshots_dir, s_name)
                    driver.save_screenshot(s_path)
                    screenshot_relative_path = f"screenshots/{s_name}"
                except Exception as e:
                    print(f"Failed to capture error screenshot: {e}")

            test_results.append({
                "id": f"TS-{len(test_results) + 1:03d}",
                "category": category,
                "component": component,
                "scenario": scenario.strip(),
                "status": "FAIL (ERROR)",
                "time": round(elapsed, 3),
                "error": error_details,
                "screenshot": screenshot_relative_path
            })
            super().addError(test, err)

    runner_result = CustomResult()
    suite.run(runner_result)
    
    total_run = len(test_results)
    passes = sum(1 for r in test_results if r["status"] == "PASS")
    failures = total_run - passes

    for r in test_results:
        if r["status"] != "PASS":
            print(f"FAILED: {r['id']} - {r['category']} - {r['component']} - {r['scenario']}")
            print(f"Error: {r['error']}")
            if r["screenshot"]:
                print(f"Screenshot: {r['screenshot']}")
            print("-" * 40)

    print("-" * 80)
    print("TEST EXECUTION COMPLETED")
    print(f"Total Executed: {total_run}")
    print(f"Passed:         {passes} (Green)")
    print(f"Failed:         {failures} (Red)")
    print("-" * 80)

    # Generate Excel Report
    print("Generating Excel report...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Execution Report"
    ws.views.sheetView[0].showGridLines = True

    # Excel Styling
    font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=10, color="333333")
    font_body_bold = Font(name="Segoe UI", size=10, bold=True, color="333333")
    
    fill_title = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    fill_header = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    fill_pass = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    fill_fail = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    
    font_pass = Font(name="Segoe UI", size=10, bold=True, color="065F46")
    font_fail = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    # Title banner
    ws.merge_cells("A1:G2")
    title_cell = ws["A1"]
    title_cell.value = "PrepWise AI - Combined Appium & Selenium Automation Test Report"
    title_cell.font = font_title
    title_cell.fill = fill_title
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for row in range(1, 3):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill_title
            cell.border = thin_border

    # Summary Statistics Block
    ws["A4"] = "Execution Date:"
    ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A5"] = "Total Test Cases:"
    ws["B5"] = total_run
    ws["D4"] = "Passed Tests:"
    ws["E4"] = passes
    ws["D5"] = "Failed Tests:"
    ws["E5"] = failures
    ws["F4"] = "Pass Rate:"
    ws["G4"] = f"{(passes/total_run)*100:.1f}%" if total_run > 0 else "0.0%"
    
    for row in [4, 5]:
        for col in [1, 2, 4, 5, 6, 7]:
            cell = ws.cell(row=row, column=col)
            if col in [1, 4, 6]:
                cell.font = font_body_bold
            else:
                cell.font = font_body
            cell.border = thin_border
            if col == 7:
                cell.font = Font(name="Segoe UI", size=11, bold=True, color="065F46" if passes == total_run else "991B1B")

    # Table Headers
    headers = ["Test ID", "Category", "Component / Module", "Test Scenario", "Status", "Execution (s)", "Error Details / Notes"]
    start_row = 7
    
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = h_text
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center" if col_idx in [1, 2, 5, 6] else "left", vertical="center")
        cell.border = thin_border

    # Table Content
    for row_idx, result in enumerate(test_results, start=start_row + 1):
        row_fill = fill_zebra if row_idx % 2 == 0 else PatternFill(fill_type=None)
        
        cell_id = ws.cell(row=row_idx, column=1, value=result["id"])
        cell_id.alignment = Alignment(horizontal="center")
        
        cell_cat = ws.cell(row=row_idx, column=2, value=result["category"])
        cell_cat.alignment = Alignment(horizontal="center")
        
        cell_comp = ws.cell(row=row_idx, column=3, value=result["component"])
        cell_scen = ws.cell(row=row_idx, column=4, value=result["scenario"])
        
        cell_stat = ws.cell(row=row_idx, column=5, value=result["status"])
        cell_stat.alignment = Alignment(horizontal="center")
        if "PASS" in result["status"]:
            cell_stat.fill = fill_pass
            cell_stat.font = font_pass
        else:
            cell_stat.fill = fill_fail
            cell_stat.font = font_fail
            
        cell_time = ws.cell(row=row_idx, column=6, value=result["time"])
        cell_time.alignment = Alignment(horizontal="center")
        
        err_val = result["error"] if result["error"] else "Successfully validated."
        if result["screenshot"]:
            err_val += f" (Screenshot captured: {result['screenshot']})"
        cell_err = ws.cell(row=row_idx, column=7, value=err_val)
        if not result["error"]:
            cell_err.font = Font(name="Segoe UI", size=10, italic=True, color="888888")

        for col_idx in range(1, 8):
            c = ws.cell(row=row_idx, column=col_idx)
            if col_idx != 5:
                c.font = font_body
            if col_idx != 5 and row_fill.fill_type:
                c.fill = row_fill
            c.border = thin_border

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 2:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    excel_output_path = os.path.join(start_dir, "test_report.xlsx")
    wb.save(excel_output_path)
    print(f"Excel report saved successfully to: {excel_output_path}")

    # Generate HTML Report using string replacement to avoid f-string JS brace collision
    print("Generating HTML report...")
    html_output_path = os.path.join(start_dir, "test_report.html")
    
    pass_rate_pct = (passes / total_run) * 100 if total_run > 0 else 0.0
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Generate HTML Table rows
    table_rows_html = ""
    for r in test_results:
        status_class = "pass-badge" if r["status"] == "PASS" else "fail-badge"
        screenshot_html = ""
        if r["screenshot"]:
            screenshot_html = f'<br/><a href="{r["screenshot"]}" target="_blank" class="screenshot-link">\ud83d\udcf1 View Failure Screenshot</a>'
        
        error_html = ""
        if r["error"]:
            error_html = f'<pre class="error-stack">{r["error"]}</pre>'

        table_rows_html += f"""
        <tr>
            <td style="text-align: center; font-weight: bold;">{r["id"]}</td>
            <td style="text-align: center;"><span class="category-badge">{r["category"]}</span></td>
            <td>{r["component"]}</td>
            <td><strong>{r["scenario"]}</strong></td>
            <td style="text-align: center;"><span class="{status_class}">{r["status"]}</span></td>
            <td style="text-align: center;">{r["time"]}s</td>
            <td>
                {error_html if r["error"] else '<span class="success-note">\u2713 Verified</span>'}
                {screenshot_html}
            </td>
        </tr>
        """

    html_template = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>PrepWise AI - Automation E2E Test Report</title>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <style>
        :root {
            --bg-color: #050505;
            --card-bg: #0f0f11;
            --border-color: #1f1f23;
            --text-primary: #ffffff;
            --text-secondary: #a1a1aa;
            --emerald-green: #10b981;
            --rose-red: #f43f5e;
            --amber-yellow: #f59e0b;
        }
        
        body {
            font-family: 'Outfit', sans-serif;
            background-color: var(--bg-color);
            color: var(--text-primary);
            margin: 0;
            padding: 2.5rem;
            line-height: 1.5;
        }

        header {
            max-width: 1200px;
            margin: 0 auto 2rem auto;
            border-bottom: 1px solid var(--border-color);
            padding-bottom: 1.5rem;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }

        .brand-title {
            font-size: 1.75rem;
            font-weight: 700;
            letter-spacing: -0.025em;
        }

        .brand-badge {
            background-color: #ffffff;
            color: #000000;
            padding: 0.1rem 0.5rem;
            font-size: 0.75rem;
            font-weight: 800;
            margin-left: 0.25rem;
            text-transform: uppercase;
        }

        .timestamp {
            font-size: 0.875rem;
            color: var(--text-secondary);
        }

        .dashboard-stats {
            max-width: 1200px;
            margin: 0 auto 2.5rem auto;
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
            gap: 1.5rem;
        }

        .stat-card {
            background-color: var(--card-bg);
            border: 1px solid var(--border-color);
            padding: 1.5rem;
            border-radius: 0px;
            position: relative;
            overflow: hidden;
        }

        .stat-card::before {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            width: 4px;
            height: 100%;
        }

        .stat-card.total::before { background-color: var(--text-secondary); }
        .stat-card.pass::before { background-color: var(--emerald-green); }
        .stat-card.fail::before { background-color: var(--rose-red); }
        .stat-card.rate::before { background-color: var(--amber-yellow); }

        .stat-label {
            font-size: 0.75rem;
            text-transform: uppercase;
            letter-spacing: 0.1em;
            color: var(--text-secondary);
            font-weight: 600;
        }

        .stat-value {
            font-size: 2.25rem;
            font-weight: 700;
            margin-top: 0.5rem;
        }

        .controls-container {
            max-width: 1200px;
            margin: 0 auto 1.5rem auto;
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 1rem;
        }

        .search-bar {
            background-color: var(--card-bg);
            border: 1px solid var(--border-color);
            color: #ffffff;
            padding: 0.75rem 1rem;
            font-family: inherit;
            font-size: 0.875rem;
            width: 300px;
            outline: none;
            transition: border-color 0.2s;
        }

        .search-bar:focus {
            border-color: #ffffff;
        }

        .filter-buttons {
            display: flex;
            gap: 0.5rem;
        }

        .filter-btn {
            background-color: transparent;
            border: 1px solid var(--border-color);
            color: var(--text-secondary);
            padding: 0.5rem 1rem;
            font-size: 0.75rem;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            cursor: pointer;
            transition: all 0.2s;
        }

        .filter-btn.active, .filter-btn:hover {
            background-color: #ffffff;
            color: #000000;
            border-color: #ffffff;
        }

        .table-container {
            max-width: 1200px;
            margin: 0 auto;
            border: 1px solid var(--border-color);
            background-color: var(--card-bg);
            overflow-x: auto;
        }

        table {
            width: 100%;
            border-collapse: collapse;
            font-size: 0.875rem;
        }

        th {
            background-color: #121215;
            color: #ffffff;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            font-size: 0.75rem;
            font-weight: 700;
            padding: 1rem;
            text-align: left;
            border-bottom: 1px solid var(--border-color);
        }

        td {
            padding: 1rem;
            border-bottom: 1px solid var(--border-color);
            vertical-align: middle;
        }

        tr:last-child td {
            border-bottom: none;
        }

        tr:hover td {
            background-color: #161619;
        }

        .category-badge {
            background-color: rgba(255,255,255,0.05);
            border: 1px solid rgba(255,255,255,0.1);
            color: #ffffff;
            padding: 0.2rem 0.5rem;
            font-size: 0.7rem;
            font-weight: 600;
            text-transform: uppercase;
            border-radius: 2px;
        }

        .pass-badge {
            background-color: rgba(16, 185, 129, 0.1);
            border: 1px solid rgba(16, 185, 129, 0.2);
            color: var(--emerald-green);
            padding: 0.25rem 0.75rem;
            font-size: 0.75rem;
            font-weight: 700;
            text-transform: uppercase;
            border-radius: 2px;
        }

        .fail-badge {
            background-color: rgba(244, 63, 94, 0.1);
            border: 1px solid rgba(244, 63, 94, 0.2);
            color: var(--rose-red);
            padding: 0.25rem 0.75rem;
            font-size: 0.75rem;
            font-weight: 700;
            text-transform: uppercase;
            border-radius: 2px;
        }

        .success-note {
            color: var(--emerald-green);
            font-style: italic;
            font-size: 0.8rem;
        }

        .screenshot-link {
            color: var(--rose-red);
            text-decoration: none;
            font-size: 0.8rem;
            font-weight: 600;
            display: inline-block;
            margin-top: 0.5rem;
            border-bottom: 1px dashed var(--rose-red);
        }

        .screenshot-link:hover {
            color: #ffffff;
            border-color: #ffffff;
        }

        .error-stack {
            background-color: #180c0e;
            border: 1px solid rgba(244, 63, 94, 0.1);
            color: #fda4af;
            padding: 0.75rem;
            font-family: monospace;
            font-size: 0.75rem;
            white-space: pre-wrap;
            word-break: break-all;
            margin: 0.5rem 0 0 0;
            max-height: 200px;
            overflow-y: auto;
        }
    </style>
</head>
<body>

    <header>
        <div>
            <div class="brand-title">PrepWise <span class="brand-badge">AI</span></div>
            <div style="font-size: 0.8rem; color: var(--emerald-green); font-weight: bold; margin-top: 0.25rem;">E2E Automation Testing Suite</div>
        </div>
        <div class="timestamp">
            Executed: <span>__NOW_STR__</span>
        </div>
    </header>

    <section class="dashboard-stats">
        <div class="stat-card total">
            <div class="stat-label">Total Executed</div>
            <div class="stat-value">__TOTAL_RUN__</div>
        </div>
        <div class="stat-card pass">
            <div class="stat-label">Passed Cases</div>
            <div class="stat-value" style="color: var(--emerald-green);">__PASSES__</div>
        </div>
        <div class="stat-card fail">
            <div class="stat-label">Failed Cases</div>
            <div class="stat-value" style="color: var(--rose-red);">__FAILURES__</div>
        </div>
        <div class="stat-card rate">
            <div class="stat-label">Success Rate</div>
            <div class="stat-value" style="color: var(--amber-yellow);">__PASS_RATE__%</div>
        </div>
    </section>

    <div class="controls-container">
        <input type="text" id="searchInput" class="search-bar" placeholder="Search scenario or module..." onkeyup="filterTable()">
        <div class="filter-buttons">
            <button class="filter-btn active" onclick="setFilter('ALL')">All</button>
            <button class="filter-btn" onclick="setFilter('PASS')">Pass</button>
            <button class="filter-btn" onclick="setFilter('FAIL')">Fail</button>
        </div>
    </div>

    <div class="table-container">
        <table id="resultsTable">
            <thead>
                <tr>
                    <th style="width: 80px; text-align: center;">ID</th>
                    <th style="width: 150px; text-align: center;">Category</th>
                    <th style="width: 200px;">Module</th>
                    <th style="width: 350px;">Scenario</th>
                    <th style="width: 100px; text-align: center;">Status</th>
                    <th style="width: 100px; text-align: center;">Duration</th>
                    <th>Execution Details</th>
                </tr>
            </thead>
            <tbody>
                __TABLE_ROWS__
            </tbody>
        </table>
    </div>

    <script>
        let currentFilter = 'ALL';

        function setFilter(filterType) {
            currentFilter = filterType;
            
            // Toggle active classes on buttons
            const buttons = document.querySelectorAll('.filter-btn');
            buttons.forEach(btn => btn.classList.remove('active'));
            
            event.target.classList.add('active');
            filterTable();
        }

        function filterTable() {
            const input = document.getElementById('searchInput');
            const filter = input.value.toLowerCase();
            const table = document.getElementById('resultsTable');
            const trs = table.getElementsByTagName('tr');

            for (let i = 1; i < trs.length; i++) {
                const tr = trs[i];
                const tds = tr.getElementsByTagName('td');
                if (tds.length === 0) continue;

                const moduleText = tds[2].textContent.toLowerCase();
                const scenarioText = tds[3].textContent.toLowerCase();
                const statusText = tds[4].textContent.trim().toUpperCase();

                const matchesSearch = moduleText.includes(filter) || scenarioText.includes(filter);
                
                let matchesStatus = true;
                if (currentFilter === 'PASS') {
                    matchesStatus = statusText === 'PASS';
                } else if (currentFilter === 'FAIL') {
                    matchesStatus = statusText.startsWith('FAIL');
                }

                if (matchesSearch && matchesStatus) {
                    tr.style.display = '';
                } else {
                    tr.style.display = 'none';
                }
            }
        }
    </script>
</body>
</html>
"""

    # Replace placeholders
    html_content = html_template \
        .replace("__NOW_STR__", now_str) \
        .replace("__TOTAL_RUN__", str(total_run)) \
        .replace("__PASSES__", str(passes)) \
        .replace("__FAILURES__", str(failures)) \
        .replace("__PASS_RATE__", f"{pass_rate_pct:.1f}") \
        .replace("__TABLE_ROWS__", table_rows_html)

    with open(html_output_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"Styled HTML report saved successfully to: {html_output_path}")
    print("=" * 80)

if __name__ == "__main__":
    main()
